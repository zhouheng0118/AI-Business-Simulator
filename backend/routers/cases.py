from __future__ import annotations

import io
from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel, Field
import database as db
from agents.playbook_generator import generate_playbook, synthesize_background

router = APIRouter(prefix="/cases", tags=["cases"])


def _excel_sheet_to_markdown(sheet) -> str:
    """Convert one openpyxl worksheet to a Markdown table, skipping blank rows."""
    rows = []
    for row in sheet.iter_rows(values_only=True):
        cells = [str(v).strip() if v is not None else "" for v in row]
        if any(cells):
            rows.append(cells)

    if not rows:
        return ""

    col_count = max(len(r) for r in rows)
    padded = [r + [""] * (col_count - len(r)) for r in rows]

    header    = "| " + " | ".join(padded[0]) + " |"
    separator = "| " + " | ".join("---" for _ in range(col_count)) + " |"
    body      = "\n".join("| " + " | ".join(r) + " |" for r in padded[1:])
    return "\n".join(filter(None, [header, separator, body]))


def _extract_excel_text(content_bytes: bytes, filename: str = "") -> tuple[str, list[str]]:
    """Parse an Excel file (.xlsx or .xls) and return (combined_markdown_text, sheet_names).

    Uses openpyxl for .xlsx and xlrd for legacy .xls files.
    """
    parts: list[str] = []
    sheet_names: list[str] = []

    if filename.lower().endswith(".xls"):
        # Legacy binary format — use xlrd
        try:
            import xlrd
        except ImportError:
            raise RuntimeError("xlrd is not installed. Run: python3.11 -m pip install xlrd")

        wb = xlrd.open_workbook(file_contents=content_bytes)
        for sheet_name in wb.sheet_names():
            ws = wb.sheet_by_name(sheet_name)
            rows = []
            for row_idx in range(ws.nrows):
                cells = []
                for val in ws.row_values(row_idx):
                    if isinstance(val, float) and val == int(val):
                        cells.append(str(int(val)))
                    else:
                        cells.append(str(val).strip() if val != "" else "")
                if any(cells):
                    rows.append(cells)

            if not rows:
                continue
            col_count = max(len(r) for r in rows)
            padded = [r + [""] * (col_count - len(r)) for r in rows]
            header    = "| " + " | ".join(padded[0]) + " |"
            separator = "| " + " | ".join("---" for _ in range(col_count)) + " |"
            body      = "\n".join("| " + " | ".join(r) + " |" for r in padded[1:])
            md = "\n".join(filter(None, [header, separator, body]))
            if md:
                parts.append(f"### {sheet_name}\n\n{md}")
                sheet_names.append(sheet_name)

    else:
        # Modern .xlsx format — use openpyxl
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is not installed. Run: python3.11 -m pip install openpyxl")

        wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            md = _excel_sheet_to_markdown(ws)
            if md:
                parts.append(f"### {sheet_name}\n\n{md}")
                sheet_names.append(sheet_name)

    return "\n\n".join(parts), sheet_names


def _table_to_markdown(table: list[list]) -> str:
    """Convert a pdfplumber table (list of rows) to a Markdown table string."""
    rows = []
    for row in table:
        cells = [str(cell).strip() if cell is not None else "" for cell in row]
        rows.append(cells)

    if not rows:
        return ""

    col_count = max(len(r) for r in rows)
    # Pad rows to equal column count
    padded = [r + [""] * (col_count - len(r)) for r in rows]

    header = "| " + " | ".join(padded[0]) + " |"
    separator = "| " + " | ".join("---" for _ in range(col_count)) + " |"
    body = "\n".join("| " + " | ".join(r) + " |" for r in padded[1:])

    parts = [header, separator]
    if body:
        parts.append(body)
    return "\n".join(parts)


def _extract_pdf_text(content_bytes: bytes) -> str:
    """Extract text and tables from a PDF using pdfplumber.

    Tables are converted to Markdown so financial data retains its structure.
    Falls back to pypdf if pdfplumber is unavailable.
    """
    try:
        import pdfplumber
    except ImportError:
        # Fallback: pypdf text-only extraction
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(content_bytes))
            return "\n\n".join(
                p.extract_text() or "" for p in reader.pages
            ).strip()
        except ImportError:
            raise RuntimeError("No PDF parsing library available (install pdfplumber).")

    page_parts: list[str] = []

    with pdfplumber.open(io.BytesIO(content_bytes)) as pdf:
        for page in pdf.pages:
            parts: list[str] = []

            # Extract tables first; track their bounding boxes to avoid
            # double-counting the same content in the plain-text pass.
            tables = page.extract_tables() or []
            table_bboxes = []
            for table_obj in page.find_tables():
                table_bboxes.append(table_obj.bbox)

            for table in tables:
                if not table:
                    continue
                md = _table_to_markdown(table)
                if md:
                    parts.append(md)

            # Extract remaining text, cropping out table regions so numbers
            # aren't repeated in both prose and table form.
            remaining_page = page
            for bbox in table_bboxes:
                try:
                    remaining_page = remaining_page.outside_bbox(bbox)
                except Exception:
                    pass

            prose = remaining_page.extract_text(x_tolerance=3, y_tolerance=3) or ""
            prose = prose.strip()
            if prose:
                parts.append(prose)

            if parts:
                page_parts.append("\n\n".join(parts))

    return "\n\n---\n\n".join(page_parts).strip()


@router.post("/parse-file")
async def parse_file(file: UploadFile = File(...)):
    """Extract plain text (and tables as Markdown) from an uploaded file."""
    filename = file.filename or ""
    content_bytes = await file.read()

    if filename.endswith((".txt", ".md")):
        try:
            text = content_bytes.decode("utf-8")
        except UnicodeDecodeError:
            text = content_bytes.decode("latin-1", errors="replace")
        return {"text": text, "file_type": "markdown" if filename.endswith(".md") else "text"}

    if filename.endswith(".pdf"):
        try:
            text = _extract_pdf_text(content_bytes)
            if not text:
                raise HTTPException(
                    status_code=422,
                    detail="PDF has no extractable text (may be scanned/image-based).",
                )
            return {"text": text, "file_type": "pdf"}
        except RuntimeError as exc:
            raise HTTPException(status_code=500, detail=str(exc))

    raise HTTPException(status_code=415, detail="Unsupported file type. Upload a .txt, .md, or .pdf file.")


@router.post("/parse-excel")
async def parse_excel(file: UploadFile = File(...)):
    """Extract financial tables from an uploaded .xlsx or .xls file."""
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=415, detail="Unsupported file type. Upload a .xlsx or .xls file.")

    content_bytes = await file.read()
    try:
        text, sheets = _extract_excel_text(content_bytes, filename)
        if not text:
            raise HTTPException(status_code=422, detail="Excel file contains no readable data.")
        return {"text": text, "sheets": sheets}
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("")
def list_cases(published_only: bool = True):
    return db.list_cases(published_only=published_only)


@router.get("/{case_id}/stats")
def get_case_stats(case_id: str):
    case = db.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    return db.get_case_stats(case_id)


@router.get("/analytics/students")
def get_all_student_analytics():
    return db.get_student_analytics()


@router.get("/{case_id}/analytics/students")
def get_case_student_analytics(case_id: str):
    case = db.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    return db.get_student_analytics(case_id)


@router.get("/{case_id}")
def get_case(case_id: str):
    case = db.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    playbook = db.get_playbook_by_case(case_id)
    return {"case": case, "playbook": playbook}


# Professor: case creation

class CreateCaseIn(BaseModel):
    title: str = Field(min_length=3)
    description: str = ""
    raw_content: str = Field(min_length=20)
    case_type: str = "decision"
    difficulty: str = "medium"
    teaching_goals: list[str] = []


@router.post("")
async def create_case_and_generate(body: CreateCaseIn):
    """Create a case and immediately generate its playbook via LLM."""
    if body.case_type not in ("decision", "analysis", "reflection"):
        raise HTTPException(status_code=422, detail="case_type must be decision, analysis, or reflection")
    if body.difficulty not in ("easy", "medium", "hard"):
        raise HTTPException(status_code=422, detail="difficulty must be easy, medium, or hard")

    case = db.create_case(
        title=body.title,
        description=body.description,
        raw_content=body.raw_content,
        case_type=body.case_type,
        difficulty=body.difficulty,
        teaching_goals=body.teaching_goals,
    )

    playbook_data = await generate_playbook(
        raw_content=body.raw_content,
        case_type=body.case_type,
        teaching_goals=body.teaching_goals,
        title=body.title,
    )

    # If the professor left description blank, populate it with the AI-generated
    # background summary so students see a rich case brief immediately.
    background_summary = playbook_data.get("background_summary", "").strip()
    if not body.description.strip() and background_summary:
        db.update_case_description(case["id"], background_summary)
        case["description"] = background_summary

    playbook = db.create_playbook(
        case_id=case["id"],
        roles=playbook_data["roles"],
        questions=playbook_data["questions"],
        info_atoms=playbook_data.get("info_atoms", []),
        checklist_items=playbook_data.get("checklist_items", []),
        calculation_challenges=playbook_data.get("calculation_challenges", []),
    )

    return {"case": case, "playbook": playbook}


class UpdateCaseIn(BaseModel):
    title: str | None = None
    description: str | None = None
    case_type: str | None = None
    difficulty: str | None = None
    teaching_goals: list[str] | None = None


@router.patch("/{case_id}")
def update_case(case_id: str, body: UpdateCaseIn):
    case = db.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    fields = {k: v for k, v in body.model_dump().items() if v is not None}
    if not fields:
        return case
    updated = db.update_case(case_id, fields)
    return updated


@router.delete("/{case_id}")
def delete_case(case_id: str):
    case = db.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    db.delete_case(case_id)
    return {"status": "deleted"}


# Professor: playbook review

class ApprovePlaybookIn(BaseModel):
    publish: bool = True


@router.post("/{case_id}/playbook/{playbook_id}/approve")
async def approve_playbook(case_id: str, playbook_id: str, body: ApprovePlaybookIn = ApprovePlaybookIn()):
    case = db.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    playbook = db.get_playbook(playbook_id)
    if not playbook or playbook["case_id"] != case_id:
        raise HTTPException(status_code=404, detail="Playbook not found")

    db.approve_playbook(playbook_id)
    if body.publish:
        db.publish_case(case_id)

    # Synthesize a student-facing background from the final basic layer atoms and
    # write it back to case.description so students see a paragraph derived from
    # the professor-reviewed info atoms rather than the raw first-pass summary.
    info_atoms = playbook.get("info_atoms") or []
    if info_atoms:
        synthesis = await synthesize_background(info_atoms)
        if synthesis:
            db.update_case_description(case_id, synthesis)

    return {"status": "approved", "case_status": "published" if body.publish else case["status"]}


class RejectPlaybookIn(BaseModel):
    notes: str = ""


@router.post("/{case_id}/playbook/{playbook_id}/reject")
def reject_playbook(case_id: str, playbook_id: str, body: RejectPlaybookIn = RejectPlaybookIn()):
    case = db.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    playbook = db.get_playbook(playbook_id)
    if not playbook or playbook["case_id"] != case_id:
        raise HTTPException(status_code=404, detail="Playbook not found")

    db.reject_playbook(playbook_id, body.notes)
    return {"status": "rejected"}


class InfoAtomItem(BaseModel):
    fact: str = Field(min_length=10)
    owner_roles: list[str] = []
    access: str = Field(pattern="^(allowed|locked)$")
    unlock_condition: str = ""
    level: int = Field(ge=0, le=3)
    category: str = ""
    objective_index: int = 0


class UpdateInfoAtomsIn(BaseModel):
    info_atoms: list[InfoAtomItem]


class UpdatePlaybookContentIn(BaseModel):
    roles: list | None = None
    questions: list | None = None
    description: str | None = None
    teaching_goals: list[str] | None = None


@router.patch("/{case_id}/playbook/{playbook_id}/content")
def update_playbook_content(case_id: str, playbook_id: str, body: UpdatePlaybookContentIn):
    case = db.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    playbook = db.get_playbook(playbook_id)
    if not playbook or playbook["case_id"] != case_id:
        raise HTTPException(status_code=404, detail="Playbook not found")
    playbook_fields: dict = {}
    if body.roles is not None:
        playbook_fields["roles"] = body.roles
    if body.questions is not None:
        playbook_fields["questions"] = body.questions
    if playbook_fields:
        db.update_playbook_content(playbook_id, playbook_fields)
    case_fields: dict = {}
    if body.description is not None:
        case_fields["description"] = body.description
    if body.teaching_goals is not None:
        case_fields["teaching_goals"] = body.teaching_goals
    if case_fields:
        db.update_case(case_id, case_fields)
    return {"status": "updated"}


@router.patch("/{case_id}/playbook/{playbook_id}/info-atoms")
def update_info_atoms(case_id: str, playbook_id: str, body: UpdateInfoAtomsIn):
    case = db.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    playbook = db.get_playbook(playbook_id)
    if not playbook or playbook["case_id"] != case_id:
        raise HTTPException(status_code=404, detail="Playbook not found")
    atoms = [item.model_dump() for item in body.info_atoms]
    db.update_playbook_info_atoms(playbook_id, atoms)
    return {"status": "updated", "count": len(atoms)}


@router.get("/{case_id}/playbook/pending")
def get_pending_playbook(case_id: str):
    case = db.get_case(case_id)
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")
    playbook = db.get_pending_playbook(case_id)
    if not playbook:
        raise HTTPException(status_code=404, detail="No playbook found for this case")
    return {"case": case, "playbook": playbook}
